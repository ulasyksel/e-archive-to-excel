from copy import copy
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


TEMPLATE_PATH = (
    Path(__file__).resolve().parent
    / "templates"
    / "excel_sablonu.xlsx"
)


def _python_value(value):
    """Pandas/NumPy değerlerini Excel'in kabul ettiği Python değerlerine çevirir."""
    if pd.isna(value):
        return None

    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass

    return value


def _copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)

    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format


def _copy_row_style(worksheet, source_row, target_row, max_column):
    for column in range(1, max_column + 1):
        _copy_cell_style(
            worksheet.cell(source_row, column),
            worksheet.cell(target_row, column),
        )

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]

    if source_dimension.height is not None:
        target_dimension.height = source_dimension.height


def _clear_values(worksheet, start_row, max_column):
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=max(worksheet.max_row, start_row),
        min_col=1,
        max_col=max_column,
    ):
        for cell in row:
            cell.value = None


def _write_dataframe(
    worksheet,
    dataframe,
    start_row,
    max_column,
    style_source_row=2,
):
    for row_offset, row_values in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=start_row,
    ):
        _copy_row_style(
            worksheet,
            source_row=style_source_row,
            target_row=row_offset,
            max_column=max_column,
        )

        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(
                row=row_offset,
                column=column_index,
                value=_python_value(value),
            )


def _multiple_payment_invoice_count(main_df):
    if main_df.empty or "FATURA NO" not in main_df.columns:
        return 0

    invoice_numbers = main_df["FATURA NO"].dropna()
    invoice_numbers = invoice_numbers[invoice_numbers.astype(str) != ""]

    if invoice_numbers.empty:
        return 0

    return int((invoice_numbers.value_counts() > 1).sum())


def build_excel(main_df, control_df, error_df, total_pdf):
    """
    Streamlit indirme düğmesinde kullanılmak üzere .xlsx dosyasını
    bellekte oluşturur ve bytes olarak döndürür.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Excel şablonu bulunamadı: {TEMPLATE_PATH}"
        )

    workbook = load_workbook(TEMPLATE_PATH)

    main_sheet = workbook["E-Fatura Özet"]
    control_sheet = workbook["Kontrol"]
    summary_sheet = workbook["Özet"]

    # Eski değerleri temizle; şablon biçimlerini koru.
    _clear_values(main_sheet, start_row=2, max_column=11)
    _clear_values(control_sheet, start_row=2, max_column=12)
    _clear_values(summary_sheet, start_row=2, max_column=2)

    # Ana tablo A:H alanına yazılır. I sütunu boş bırakılır.
    # J ve K sütunları yardımcı prefix/mağaza alanlarıdır ve gizlidir.
    main_export_df = main_df.copy()

    _write_dataframe(
        main_sheet,
        main_export_df,
        start_row=2,
        max_column=11,
        style_source_row=2,
    )

    for row_number in range(2, len(main_export_df) + 2):
        invoice_no = main_sheet.cell(row_number, 2).value
        store = main_sheet.cell(row_number, 7).value

        main_sheet.cell(row_number, 9).value = None
        main_sheet.cell(row_number, 10).value = (
            str(invoice_no)[:3] if invoice_no else None
        )
        main_sheet.cell(row_number, 11).value = store

    main_sheet.column_dimensions["J"].hidden = True
    main_sheet.column_dimensions["K"].hidden = True

    # Sayısal sütunların Excel sayı biçimi.
    for row_number in range(2, len(main_export_df) + 2):
        for column_number in (3, 4, 5):
            main_sheet.cell(
                row=row_number,
                column=column_number,
            ).number_format = "#,##0.00"

    _write_dataframe(
        control_sheet,
        control_df,
        start_row=2,
        max_column=12,
        style_source_row=2,
    )

    for row_number in range(2, len(control_df) + 2):
        for column_number in range(4, 11):
            control_sheet.cell(
                row=row_number,
                column=column_number,
            ).number_format = "#,##0.00"

    included_document_count = len(control_df)
    control_required_count = (
        int((control_df["DURUM"] == "KONTROL").sum())
        if not control_df.empty and "DURUM" in control_df.columns
        else 0
    )

    metrics = [
        ("Toplam PDF", int(total_pdf)),
        ("Ana tabloya dahil edilen belge", included_document_count),
        ("Normal e-Fatura/e-Arşiv belge", included_document_count),
        ("Global Blue Tax Free Form", 0),
        ("Toplam ödeme satırı", len(main_df)),
        (
            "Birden fazla ödeme içeren fatura",
            _multiple_payment_invoice_count(main_df),
        ),
        ("Kontrol gereken belge", control_required_count),
    ]

    # Metrikler.
    current_row = 2

    for metric_name, metric_value in metrics:
        _copy_row_style(
            summary_sheet,
            source_row=2,
            target_row=current_row,
            max_column=2,
        )
        summary_sheet.cell(current_row, 1).value = metric_name
        summary_sheet.cell(current_row, 2).value = metric_value
        current_row += 1

    current_row += 1

    # Ödeme türü özet başlığı.
    _copy_cell_style(
        summary_sheet["A1"],
        summary_sheet.cell(current_row, 1),
    )
    _copy_cell_style(
        summary_sheet["B1"],
        summary_sheet.cell(current_row, 2),
    )
    summary_sheet.cell(current_row, 1).value = "ÖDEME ŞEKLİ"
    summary_sheet.cell(current_row, 2).value = "SATIR SAYISI"
    current_row += 1

    if not main_df.empty and "ÖDEME ŞEKLİ" in main_df.columns:
        payment_counts = (
            main_df["ÖDEME ŞEKLİ"]
            .fillna("DİĞER/BELİRSİZ")
            .value_counts()
            .sort_index()
        )

        for payment_method, row_count in payment_counts.items():
            _copy_row_style(
                summary_sheet,
                source_row=2,
                target_row=current_row,
                max_column=2,
            )
            summary_sheet.cell(current_row, 1).value = str(payment_method)
            summary_sheet.cell(current_row, 2).value = int(row_count)
            current_row += 1

    # Hatalı dosyalar uygulamadaki "Okuma Hataları" sekmesinde gösterilir.

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()
